#!/usr/bin/env python3
"""
创建合并的Excel文件，包含所有算法的性能数据
"""

import csv
import pandas as pd

def create_combined_excel():
    """
    创建包含所有算法性能数据的Excel文件
    """
    print("=== 创建合并的Excel文件 ===")
    
    # 原始表格数据 (DFS, BFS, HGBPC)
    original_data = [
        [1, 4, 4, 2, 0, 1.465, 0.817, 1.480, 0.001],
        [2, 4, 4, 2, 1, 1.413, 0.800, 1.554, 0.001],
        [3, 4, 4, 3, 0, 1.321, 0.763, 1.299, 0.001],
        [4, 4, 4, 3, 1, 1.155, 0.701, 1.100, 0.001],
        [5, 4, 4, 4, 0, 1.194, 0.715, 1.190, 0.001],
        [6, 4, 4, 4, 1, 1.032, 0.656, 0.973, 0.001],
        [7, 5, 4, 2, 0, 32.816, 4.538, 10.100, 0.001],
        [8, 5, 4, 2, 1, 30.757, 4.467, 8.330, 0.001],
        [9, 5, 4, 3, 0, 28.234, 4.442, 8.000, 0.001],
        [10, 5, 4, 3, 1, 26.687, 4.281, 7.597, 0.001],
        [11, 5, 4, 4, 0, 27.444, 4.355, 7.695, 0.001],
        [12, 5, 4, 4, 1, 22.959, 4.155, 7.150, 0.001],
        [13, 6, 4, 2, 0, 805.379, 23.437, 47.341, 0.001],
        [14, 6, 4, 2, 1, 775.239, 23.530, 44.858, 0.001],
        [15, 6, 4, 3, 0, 776.429, 23.778, 43.548, 0.001],
        [16, 6, 4, 3, 1, 747.787, 23.069, 42.574, 0.001],
        [17, 6, 4, 4, 0, 756.214, 23.210, 44.042, 0.001],
        [18, 6, 4, 4, 1, 8347.513, 23.786, 42.039, 0.001],
    ]
    
    # 读取A*和双向BFS数据
    astar_bidirectional_data = {}
    try:
        with open('rq1_average_times_analysis.csv', 'r') as csvfile:
            reader = csv.DictReader(csvfile)
            for row in reader:
                no = int(row['No'])
                astar_time = float(row['astar_avg_time_ms'])
                bidirectional_time = float(row['bidirectional_bfs_avg_time_ms'])
                astar_bidirectional_data[no] = (astar_time, bidirectional_time)
    except FileNotFoundError:
        print("错误: 找不到 rq1_average_times_analysis.csv 文件")
        return
    
    # 合并数据
    combined_data = []
    for row in original_data:
        no = row[0]
        n, k, r, h = row[1:5]
        dfs_time, bfs_time, hgbpc_uf_build, hgbpc_uf_query = row[5:9]
        
        # 获取A*和双向BFS数据
        if no in astar_bidirectional_data:
            astar_time, bidirectional_time = astar_bidirectional_data[no]
        else:
            astar_time, bidirectional_time = 0, 0
            print(f"警告: 配置 {no} 的A*和双向BFS数据缺失")
        
        combined_data.append([
            no, n, k, r, h,
            dfs_time, bfs_time, astar_time, bidirectional_time,
            hgbpc_uf_build, hgbpc_uf_query
        ])
    
    # 计算平均值
    avg_dfs = sum(row[5] for row in combined_data) / len(combined_data)
    avg_bfs = sum(row[6] for row in combined_data) / len(combined_data)
    avg_astar = sum(row[7] for row in combined_data) / len(combined_data)
    avg_bidirectional = sum(row[8] for row in combined_data) / len(combined_data)
    avg_hgbpc_build = sum(row[9] for row in combined_data) / len(combined_data)
    avg_hgbpc_query = sum(row[10] for row in combined_data) / len(combined_data)
    
    # 添加平均值行
    combined_data.append([
        "Avg", "-", "-", "-", "-",
        avg_dfs, avg_bfs, avg_astar, avg_bidirectional,
        avg_hgbpc_build, avg_hgbpc_query
    ])
    
    # 创建DataFrame
    columns = [
        'No', 'n', 'k', 'r', 'h',
        'DFS_search_time', 'BFS_search_time', 'A*_search_time', 'Bidirectional_BFS_search_time',
        'HGBPC_uf_build_time', 'HGBPC_uf_query_time'
    ]
    
    df = pd.DataFrame(combined_data, columns=columns)
    
    # 保存为Excel文件
    excel_filename = 'combined_algorithm_performance.xlsx'
    
    with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
        # 主数据表
        df.to_excel(writer, sheet_name='Algorithm_Performance', index=False)
        
        # 创建格式化的表格
        workbook = writer.book
        worksheet = writer.sheets['Algorithm_Performance']
        
        # 设置列宽
        column_widths = [5, 5, 5, 5, 5, 15, 15, 15, 20, 18, 18]
        for i, width in enumerate(column_widths):
            worksheet.column_dimensions[chr(65 + i)].width = width
        
        # 添加标题行格式
        from openpyxl.styles import Font, PatternFill, Alignment
        
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        center_alignment = Alignment(horizontal="center", vertical="center")
        
        for col in range(1, len(columns) + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
        
        # 设置数据格式
        for row in range(2, len(combined_data) + 2):
            for col in range(1, len(columns) + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.alignment = center_alignment
                
                # 对时间数据保留3位小数
                if col > 5:  # 时间列
                    if isinstance(cell.value, (int, float)) and cell.value != 0:
                        cell.number_format = '0.000'
        
        # 高亮平均值行
        avg_row = len(combined_data) + 1
        avg_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        for col in range(1, len(columns) + 1):
            cell = worksheet.cell(row=avg_row, column=col)
            cell.fill = avg_fill
            cell.font = Font(bold=True)
    
    print(f"Excel文件已创建: {excel_filename}")
    
    # 打印汇总信息
    print("\n=== 数据汇总 ===")
    print(f"总配置数: {len(combined_data) - 1}")  # 减去平均值行
    print(f"DFS平均时间: {avg_dfs:.3f} ms")
    print(f"BFS平均时间: {avg_bfs:.3f} ms")
    print(f"A*平均时间: {avg_astar:.3f} ms")
    print(f"双向BFS平均时间: {avg_bidirectional:.3f} ms")
    print(f"HGBPC构建时间: {avg_hgbpc_build:.3f} ms")
    print(f"HGBPC查询时间: {avg_hgbpc_query:.3f} ms")
    
    # 性能比较
    print("\n=== 性能比较 ===")
    print(f"双向BFS vs BFS: {avg_bidirectional/avg_bfs:.2f}x")
    print(f"A* vs BFS: {avg_astar/avg_bfs:.2f}x")
    print(f"双向BFS vs A*: {avg_bidirectional/avg_astar:.2f}x")
    
    return df

def create_latex_table(df):
    """
    创建完整的LaTeX表格
    """
    print("\n=== 完整LaTeX表格 ===")
    print("\\begin{table}")
    print("\\centering")
    print("\\caption{Time (ms) Performance of BFS, DFS, A*, BBFS, and HGBPC with Source and Target Fault-free Nodes from Different Components}")
    print("\\label{tab:combined_algorithm_performance}")
    print("\\begin{tabular}{c|c|c|c|c|c|c|c|c|cc}")
    print("\\hline")
    print("\\multirow{2}{*}{\\textbf{No}} & \\multirow{2}{*}{\\textbf{n}} & \\multirow{2}{*}{\\textbf{k}} & \\multirow{2}{*}{\\textbf{r}} & \\multirow{2}{*}{\\textbf{h}} & \\textbf{DFS} & \\textbf{BFS} & \\textbf{A*} & \\textbf{BBFS} & \\multicolumn{2}{c}{\\textbf{HGBPC}} \\\\")
    print("\\cline{6-11}")
    print(" & & & & & \\textbf{search\\_time} & \\textbf{search\\_time} & \\textbf{search\\_time} & \\textbf{search\\_time} & \\textbf{uf\\_build\\_time} & \\textbf{uf\\_query\\_time} \\\\")
    print("\\hline")
    
    for _, row in df.iterrows():
        if row['No'] == "Avg":
            print("\\hline")
            print(f"\\textbf{{Avg}} & - & - & - & - & \\textbf{{{row['DFS_search_time']:.3f}}} & \\textbf{{{row['BFS_search_time']:.3f}}} & \\textbf{{{row['A*_search_time']:.3f}}} & \\textbf{{{row['Bidirectional_BFS_search_time']:.3f}}} & \\textbf{{{row['HGBPC_uf_build_time']:.3f}}} & \\textbf{{{row['HGBPC_uf_query_time']:.3f}}} \\\\")
        else:
            print(f"\\textbf{{{int(row['No'])}}} & {int(row['n'])} & {int(row['k'])} & {int(row['r'])} & {int(row['h'])} & {row['DFS_search_time']:.3f} & {row['BFS_search_time']:.3f} & {row['A*_search_time']:.3f} & {row['Bidirectional_BFS_search_time']:.3f} & {row['HGBPC_uf_build_time']:.3f} & {row['HGBPC_uf_query_time']:.3f} \\\\")
    
    print("\\hline")
    print("\\end{tabular}")
    print("\\end{table}")

if __name__ == "__main__":
    try:
        df = create_combined_excel()
        if df is not None:
            create_latex_table(df)
    except ImportError:
        print("错误: 需要安装 pandas 和 openpyxl")
        print("请运行: pip install pandas openpyxl")
    except Exception as e:
        print(f"创建Excel文件时出错: {e}")
