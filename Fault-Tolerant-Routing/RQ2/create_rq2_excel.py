#!/usr/bin/env python3
"""
创建RQ2的Excel文件（如果pandas可用）
"""

try:
    import pandas as pd
    from openpyxl.styles import Font, PatternFill, Alignment
    
    def create_excel_from_csv():
        """
        从CSV文件创建格式化的Excel文件
        """
        print("=== 创建RQ2 Excel文件 ===")
        
        # 读取CSV文件
        df = pd.read_csv('rq2_combined_algorithm_performance.csv')
        
        # 创建Excel文件
        excel_filename = 'rq2_combined_algorithm_performance.xlsx'
        
        with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
            # 写入主数据表
            df.to_excel(writer, sheet_name='RQ2_Algorithm_Performance', index=False)
            
            # 获取工作簿和工作表
            workbook = writer.book
            worksheet = writer.sheets['RQ2_Algorithm_Performance']
            
            # 设置列宽
            column_widths = [5, 5, 5, 5, 5, 12, 12, 12, 12, 12, 12, 18, 18, 18, 18, 12, 12]
            for i, width in enumerate(column_widths):
                if i < len(df.columns):
                    worksheet.column_dimensions[chr(65 + i)].width = width
            
            # 设置标题行格式
            header_font = Font(bold=True, size=11)
            header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            center_alignment = Alignment(horizontal="center", vertical="center")
            
            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
            
            # 设置数据格式
            for row in range(2, len(df) + 2):
                for col in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell.alignment = center_alignment
                    
                    # 对时间和路径数据设置格式
                    if col > 5:  # 数值列
                        if isinstance(cell.value, (int, float)) and cell.value != 0:
                            if 'time' in df.columns[col-1].lower():
                                cell.number_format = '0.000'  # 时间保留3位小数
                            elif 'length' in df.columns[col-1].lower():
                                cell.number_format = '0.00'   # 路径长度保留2位小数
            
            # 高亮平均值行
            avg_row = len(df) + 1
            avg_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=avg_row, column=col)
                cell.fill = avg_fill
                cell.font = Font(bold=True)
        
        print(f"Excel文件已创建: {excel_filename}")
        
        # 打印一些统计信息
        avg_row_data = df[df['No'] == 'Avg'].iloc[0]
        print(f"\n=== 关键性能指标 ===")
        print(f"A*平均时间: {avg_row_data['A*_path_time_ms']:.3f} ms")
        print(f"双向BFS平均时间: {avg_row_data['Bidirectional_BFS_path_time_ms']:.3f} ms")
        print(f"BFS平均时间: {avg_row_data['BFS_path_time_ms']:.3f} ms")
        print(f"A*平均路径长度: {avg_row_data['A*_path_length']:.2f}")
        print(f"双向BFS平均路径长度: {avg_row_data['Bidirectional_BFS_path_length']:.2f}")
        print(f"BFS平均路径长度: {avg_row_data['BFS_path_length']:.2f}")
        
        print(f"\n=== 性能优势 ===")
        astar_vs_bfs_time = avg_row_data['A*_path_time_ms'] / avg_row_data['BFS_path_time_ms']
        bidirectional_vs_bfs_time = avg_row_data['Bidirectional_BFS_path_time_ms'] / avg_row_data['BFS_path_time_ms']
        astar_vs_bidirectional_time = avg_row_data['A*_path_time_ms'] / avg_row_data['Bidirectional_BFS_path_time_ms']
        
        print(f"A*比BFS快: {1/astar_vs_bfs_time:.1f}倍 ({astar_vs_bfs_time:.3f}x)")
        print(f"双向BFS比BFS快: {1/bidirectional_vs_bfs_time:.1f}倍 ({bidirectional_vs_bfs_time:.3f}x)")
        print(f"A*比双向BFS快: {1/astar_vs_bidirectional_time:.1f}倍 ({astar_vs_bidirectional_time:.3f}x)")

    if __name__ == "__main__":
        create_excel_from_csv()

except ImportError:
    print("pandas或openpyxl未安装，无法创建Excel文件")
    print("CSV文件已可用: rq2_combined_algorithm_performance.csv")
except Exception as e:
    print(f"创建Excel文件时出错: {e}")
    print("CSV文件已可用: rq2_combined_algorithm_performance.csv")
